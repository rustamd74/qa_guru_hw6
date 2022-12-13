from zipfile import ZipFile, ZIP_DEFLATED
import os
from PyPDF2 import PdfReader
from openpyxl.reader.excel import load_workbook

current_dir = os.path.dirname(os.path.abspath(__file__))
path_resources = os.path.join(current_dir, 'resources')
path_myzip = os.path.join(path_resources, 'archive.zip')


# Архив

def test_create_archive():
    with ZipFile(os.path.join(path_resources, 'archive.zip'), 'w', compression=ZIP_DEFLATED) as myzip:
        myzip.write(os.path.join(path_resources, 'names.csv'), 'names.csv')
        myzip.write(os.path.join(path_resources, 'docs-pytest-org-en-latest.pdf'), 'docs-pytest-org-en-latest.pdf')
        myzip.write(os.path.join(path_resources, 'accounts.xlsx'), 'accounts.xlsx')


# Проверка csv-файла
def test_csv():
    with ZipFile(path_myzip, 'r') as zip_file:
        csvfile = str(zip_file.read('names.csv'))
        assert csvfile.__contains__("Anna")

# Проверка pdf-файла
def test_pdf():
    with ZipFile(path_myzip, 'r') as zip_pdf_file:
        pdf_file = zip_pdf_file.extract('docs-pytest-org-en-latest.pdf')
        reader = PdfReader(pdf_file)
        page = reader.pages[0]
        text = page.extractText()
        assert 'Release 0.1' in text
        os.remove(pdf_file)

# Проверка xlsx-файла
def test_xls():
    with ZipFile(path_myzip, 'r') as zip_xlsx_file:
        xlsx_file = zip_xlsx_file.extract('accounts.xlsx')
        xlsxfile = load_workbook(xlsx_file)
        sheet = xlsxfile.active
        assert sheet.cell(row=3, column=3).value == 'Smith'
        os.remove(xlsx_file)

